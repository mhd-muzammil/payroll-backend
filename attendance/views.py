import logging
import math
import openpyxl
import re
from datetime import datetime, time, date as dt_date, timedelta
from django.db import transaction
from django.utils import timezone
from django.utils.timezone import make_aware
from django.utils.dateparse import parse_date
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Attendance, LeaveRequest
from .serializer import AttendanceSerializer, LeaveRequestSerializer
from employees.models import Employee
from authentication.models import User, get_allowed_branches

logger = logging.getLogger(__name__)

def haversine_distance(lat1, lon1, lat2, lon2):
    # approximate radius of earth in km
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) * math.sin(dlat / 2) + math.cos(math.radians(lat1)) \
        * math.cos(math.radians(lat2)) * math.sin(dlon / 2) * math.sin(dlon / 2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c * 1000  # returns distance in meters

class AttendanceViewSet(viewsets.ModelViewSet):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["post"])
    def import_sheet(self, request):
        user = request.user
        role = "superadmin" if user.is_superuser else getattr(user, 'role', 'employee')
        if role not in ["superadmin", "admin"]:
            return Response({"detail": "Permission denied."}, status=403)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"detail": "No file uploaded."}, status=400)

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception as e:
            return Response({"detail": f"Failed to parse Excel file: {str(e)}"}, status=400)

        def normalize_val(val):
            if val is None:
                return ""
            return " ".join(str(val).split()).strip().lower()

        # Helper to parse dates in header
        def parse_header_date(val, default_year, default_month):
            if not val:
                return None
            if isinstance(val, (datetime, dt_date)):
                d = val.date() if hasattr(val, 'date') else val
                if d.year in [1899, 1900]:
                    day_val = d.day
                    if d.year == 1899 and d.month == 12 and d.day == 31:
                        day_val = 1
                    try:
                        return dt_date(default_year, default_month, day_val)
                    except ValueError:
                        pass
                return d
            if isinstance(val, (int, float)):
                try:
                    day_val = int(val)
                    if 1 <= day_val <= 31:
                        return dt_date(default_year, default_month, day_val)
                except (ValueError, TypeError):
                    pass
            val_str = normalize_val(val)
            if val_str.isdigit():
                try:
                    day_val = int(val_str)
                    if 1 <= day_val <= 31:
                        return dt_date(default_year, default_month, day_val)
                except ValueError:
                    pass
            # Try to search for DD-MMM pattern anywhere in the normalized string
            # e.g. "25-mar", "25-march", "wed 25-mar"
            match = re.search(r'\b(\d{1,2})[-/]([A-Za-z]+)\b', val_str)
            if match:
                day = int(match.group(1))
                month_name = match.group(2)
                try:
                    dt = datetime.strptime(f"{day}-{month_name[:3]}-{default_year}", "%d-%b-%Y")
                    return dt.date()
                except ValueError:
                    pass
            # Pattern: DD-MM-YYYY or YYYY-MM-DD anywhere
            match_numeric = re.search(r'\b(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})\b', val_str)
            if match_numeric:
                part1 = match_numeric.group(1)
                part2 = match_numeric.group(2)
                part3 = match_numeric.group(3)
                try:
                    dt = datetime.strptime(f"{part1}-{part2}-{part3}", "%d-%m-%Y")
                    return dt.date()
                except ValueError:
                    pass
                try:
                    dt = datetime.strptime(f"{part1}-{part2}-{part3}", "%m-%d-%Y")
                    return dt.date()
                except ValueError:
                    pass
            return None

        # Helper to parse time string/floats
        def parse_excel_time(val):
            if val is None:
                return None
            if isinstance(val, time):
                return val
            if isinstance(val, datetime):
                return val.time()
            
            val_str = str(val).strip()
            if val_str == "" or val_str == "00:00" or val_str == "00:00:00":
                return None
                
            for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M", "%H:%M:%S", "%I:%M:%S %p", "%I:%M:%S%p"):
                try:
                    return datetime.strptime(val_str, fmt).time()
                except ValueError:
                    pass
                    
            try:
                val_float = float(val)
                total_seconds = int(round(val_float * 86400))
                hour = (total_seconds // 3600) % 24
                minute = (total_seconds % 3600) // 60
                second = total_seconds % 60
                return time(hour, minute, second)
            except ValueError:
                pass

            return None

        # Map a weekday label cell ("sat", "Sun", "MON", "fri.") to a Python
        # weekday number (Mon=0 .. Sun=6), or None if it isn't a weekday label.
        WEEKDAY_NAMES = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}
        def weekday_label(val):
            if not isinstance(val, str):
                return None
            key = re.sub(r'[^a-z]', '', val.strip().lower())[:3]
            return WEEKDAY_NAMES.get(key)

        # Scan the first sheet to locate global "Year" or "Month" default
        default_year = datetime.now().year
        default_month = datetime.now().month
        if wb.worksheets:
            first_sheet = wb.worksheets[0]
            first_sheet_rows = []
            for row in first_sheet.iter_rows(values_only=True):
                if len(first_sheet_rows) >= 6:
                    break
                first_sheet_rows.append(row)
            num_r = len(first_sheet_rows)
            num_c = len(first_sheet_rows[0]) if num_r > 0 else 0
            for r in range(1, min(6, num_r + 1)):
                for c in range(1, num_c + 1):
                    val = first_sheet_rows[r-1][c-1]
                    val_norm = normalize_val(val)
                    if "year" in val_norm:
                        if ":" in val_norm:
                            try:
                                default_year = int(val_norm.split(":")[1].strip())
                            except ValueError:
                                pass
                        if default_year == datetime.now().year and c < num_c:
                            right_val = first_sheet_rows[r-1][c]
                            if right_val:
                                try:
                                    default_year = int(str(right_val).strip())
                                except ValueError:
                                    pass
                    if "month" in val_norm:
                        if ":" in val_norm:
                            try:
                                default_month = int(val_norm.split(":")[1].strip())
                            except ValueError:
                                pass
                        if default_month == datetime.now().month and c < num_c:
                            right_val = first_sheet_rows[r-1][c]
                            if right_val:
                                try:
                                    default_month = int(str(right_val).strip())
                                except ValueError:
                                    pass

        employee_data = {}

        for sheet in wb.worksheets:
            # Determine branch name from sheet title
            sheet_title_norm = " ".join(sheet.title.split()).strip().lower()
            sheet_branch = None
            if "chennai" in sheet_title_norm:
                sheet_branch = "Chennai"
            elif "vellore" in sheet_title_norm:
                sheet_branch = "Vellore"
            elif "salem" in sheet_title_norm:
                sheet_branch = "Salem"
            elif "kanchipuram" in sheet_title_norm or "kanchi" in sheet_title_norm:
                sheet_branch = "Kanchipuram"
            elif "hosur" in sheet_title_norm:
                sheet_branch = "Hosur"

            # Only branch sheets carry per-person attendance. Other tabs in these
            # workbooks are non-attendance and MUST NOT be imported:
            #   - "MASTER": a blank salary/roster template (no codes or names).
            #   - "House Keeping": a per-branch roll-up whose "names" are branch
            #     names ("Chennai", "Vellore", ...), not people.
            # Importing them would invent bogus employees, so skip any sheet whose
            # title doesn't map to a known branch.
            if sheet_branch is None:
                logger.info("Excel import: skipping non-branch sheet '%s'", sheet.title)
                continue

            # Efficiently read rows up to 100 consecutive completely empty rows
            sheet_values = []
            empty_streak = 0
            for row in sheet.iter_rows(values_only=True):
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    empty_streak += 1
                    if empty_streak > 100:
                        break
                else:
                    if empty_streak > 0:
                        num_cells = len(row)
                        for _ in range(empty_streak):
                            sheet_values.append((None,) * num_cells)
                        empty_streak = 0
                    sheet_values.append(row)

            if not sheet_values:
                continue

            num_rows = len(sheet_values)
            num_cols = len(sheet_values[0]) if num_rows > 0 else 0

            # Scan first 5 rows of this sheet to locate sheet-specific "Year" or "Month"
            sheet_year = None
            sheet_month = None
            for r in range(1, min(6, num_rows + 1)):
                for c in range(1, num_cols + 1):
                    val = sheet_values[r-1][c-1]
                    val_norm = normalize_val(val)
                    if "year" in val_norm:
                        if ":" in val_norm:
                            try:
                                sheet_year = int(val_norm.split(":")[1].strip())
                            except ValueError:
                                pass
                        if not sheet_year and c < num_cols:
                            right_val = sheet_values[r-1][c]
                            if right_val:
                                try:
                                    sheet_year = int(str(right_val).strip())
                                except ValueError:
                                    pass
                    if "month" in val_norm:
                        if ":" in val_norm:
                            try:
                                sheet_month = int(val_norm.split(":")[1].strip())
                            except ValueError:
                                pass
                        if not sheet_month and c < num_cols:
                            right_val = sheet_values[r-1][c]
                            if right_val:
                                try:
                                    sheet_month = int(str(right_val).strip())
                                except ValueError:
                                    pass

            if not sheet_year:
                sheet_year = default_year
            if not sheet_month:
                sheet_month = default_month

            # Find header row containing Employee Code and Name
            header_row_idx = None
            emp_code_col_idx = None
            emp_name_col_idx = None
            branch_col_idx = None

            for r in range(1, min(21, num_rows + 1)):
                temp_code_col = None
                temp_name_col = None
                temp_branch_col = None
                for c in range(1, num_cols + 1):
                    val = sheet_values[r-1][c-1]
                    val_norm = normalize_val(val)
                    
                    is_code_match = ("emp" in val_norm and "code" in val_norm) or (val_norm == "code") or ("employee" in val_norm and "code" in val_norm)
                    is_name_match = ("employee" in val_norm and "name" in val_norm) or ("emp" in val_norm and "name" in val_norm) or (val_norm == "name")
                    is_branch_match = ("branch" in val_norm) or ("region" in val_norm) or ("location" in val_norm) or ("city" in val_norm) or ("zone" in val_norm)
                    
                    if is_code_match:
                        temp_code_col = c
                    elif is_name_match:
                        temp_name_col = c
                    elif is_branch_match:
                        temp_branch_col = c
                
                if temp_code_col and temp_name_col:
                    header_row_idx = r
                    emp_code_col_idx = temp_code_col
                    emp_name_col_idx = temp_name_col
                    branch_col_idx = temp_branch_col
                    break

            if not header_row_idx:
                # Fallback: find them independently
                for r in range(1, min(21, num_rows + 1)):
                    for c in range(1, num_cols + 1):
                        val = sheet_values[r-1][c-1]
                        val_norm = normalize_val(val)
                        if (not emp_code_col_idx) and (("emp" in val_norm and "code" in val_norm) or (val_norm == "code") or ("employee" in val_norm and "code" in val_norm)):
                            emp_code_col_idx = c
                            header_row_idx = r
                        elif (not emp_name_col_idx) and (("employee" in val_norm and "name" in val_norm) or ("emp" in val_norm and "name" in val_norm) or (val_norm == "name")):
                            emp_name_col_idx = c
                            if not header_row_idx:
                                header_row_idx = r
                        elif (not branch_col_idx) and (("branch" in val_norm) or ("region" in val_norm) or ("location" in val_norm) or ("city" in val_norm) or ("zone" in val_norm)):
                            branch_col_idx = c

            if header_row_idx and not branch_col_idx:
                for c in range(1, num_cols + 1):
                    val = sheet_values[header_row_idx-1][c-1]
                    val_norm = normalize_val(val)
                    if ("branch" in val_norm) or ("region" in val_norm) or ("location" in val_norm) or ("city" in val_norm) or ("zone" in val_norm):
                        branch_col_idx = c
                        break

            if not header_row_idx or not emp_name_col_idx or not emp_code_col_idx:
                logger.warning("Excel import: could not identify header row in sheet '%s', skipping", sheet.title)
                continue

            # Find type/status column index
            type_col_idx = None
            for r in range(header_row_idx + 1, min(header_row_idx + 12, num_rows + 1)):
                for c in range(1, num_cols + 1):
                    if c == branch_col_idx:
                        continue
                    val = sheet_values[r-1][c-1]
                    val_norm = normalize_val(val)
                    if val_norm in ["in time", "out time", "actual hrs", "present day", "present days"]:
                        type_col_idx = c
                        break
                if type_col_idx:
                    break

            if not type_col_idx:
                type_col_idx = emp_name_col_idx + 1

            # Map date columns and determine header bounds
            date_cols = {}
            dates_on_header_row = 0
            dates_on_below_row = 0
            
            for c in range(1, num_cols + 1):
                if c in [emp_code_col_idx, emp_name_col_idx, type_col_idx, branch_col_idx]:
                    continue
                
                val_header = sheet_values[header_row_idx - 1][c - 1]
                val_below = sheet_values[header_row_idx][c - 1] if header_row_idx < num_rows else None
                
                parsed_header = parse_header_date(val_header, sheet_year, sheet_month)
                parsed_below = parse_header_date(val_below, sheet_year, sheet_month)
                
                if parsed_header:
                    dates_on_header_row += 1
                if parsed_below:
                    dates_on_below_row += 1
                    
                parsed_date = parsed_header or parsed_below
                if parsed_date:
                    date_cols[c] = parsed_date

            if not date_cols:
                logger.warning("Excel import: no valid date columns found in sheet '%s', skipping", sheet.title)
                continue

            if dates_on_below_row > dates_on_header_row:
                start_data_row = header_row_idx + 2
            else:
                start_data_row = header_row_idx + 1

            # --- Correct date columns using the weekday-label row as ground truth ---
            # Two real-world defects in these biometric exports are repaired here:
            #  (a) Wrong YEAR: cells stamped e.g. 2025 whose weekday labels only fit
            #      2026 (a year/month-boundary glitch). Fixed PER COLUMN by snapping
            #      the date's year to the one near the sheet's year whose weekday
            #      matches that column's label.
            #  (b) Garbage columns: trailing "=prev+1" formulas that collapsed to
            #      1899/1900 epoch dates (which parse_header_date then remaps onto a
            #      bogus day of the sheet's month, colliding with real days). A
            #      column whose label CANNOT be satisfied by any nearby year is
            #      treated as broken and rebuilt from the running daily sequence
            #      (previous trustworthy day + 1) — its own date and stale label are
            #      both ignored.
            ordered_cols = sorted(date_cols.keys())
            date_value_row = header_row_idx + 1 if start_data_row == header_row_idx + 2 else header_row_idx

            best_label_row = None
            best_label_count = 0
            for cand_row in (header_row_idx, header_row_idx - 1, header_row_idx + 1):
                if cand_row < 1 or cand_row > num_rows or cand_row == date_value_row:
                    continue
                cnt = sum(1 for c in ordered_cols if weekday_label(sheet_values[cand_row - 1][c - 1]) is not None)
                if cnt > best_label_count:
                    best_label_count = cnt
                    best_label_row = cand_row

            if best_label_row and best_label_count >= max(3, len(ordered_cols) // 2):
                labels = {c: weekday_label(sheet_values[best_label_row - 1][c - 1]) for c in ordered_cols}
                year_window = sorted({sheet_year, sheet_year - 1, sheet_year + 1},
                                     key=lambda y: abs(y - sheet_year))
                corrections = 0
                broken = set()

                # (a) Snap each labelled column's year to match its weekday label.
                #     Columns whose label no nearby year can satisfy are BROKEN.
                for c in ordered_cols:
                    lbl = labels.get(c)
                    if lbl is None:
                        continue
                    d = date_cols[c]
                    if d.weekday() == lbl:
                        continue
                    snapped = None
                    for y in year_window:
                        try:
                            cand = d.replace(year=y)
                        except ValueError:
                            continue
                        if cand.weekday() == lbl:
                            snapped = cand
                            break
                    if snapped is not None:
                        date_cols[c] = snapped
                        corrections += 1
                    else:
                        broken.add(c)

                # (b) Rebuild broken columns as a running daily sequence: forward
                #     from the previous trustworthy day, then a backward pass for any
                #     leading broken columns that had no good day to their left.
                fixed = set()
                last_good = None
                for c in ordered_cols:
                    if c not in broken:
                        last_good = date_cols[c]
                    elif last_good is not None:
                        last_good = last_good + timedelta(days=1)
                        date_cols[c] = last_good
                        fixed.add(c)
                        corrections += 1
                next_good = None
                for c in reversed(ordered_cols):
                    if c not in broken or c in fixed:
                        next_good = date_cols[c]
                    elif next_good is not None:
                        next_good = next_good - timedelta(days=1)
                        date_cols[c] = next_good
                        corrections += 1

                if corrections:
                    logger.info("Excel import '%s': corrected %d date column(s) via weekday labels (%s .. %s)",
                                sheet.title, corrections, date_cols[ordered_cols[0]], date_cols[ordered_cols[-1]])

            # (c) Final contiguity pass. The date columns sit left-to-right in
            # calendar order, one per day, so each should be exactly one day
            # after the previous. A stale "next cycle" block left stamped with
            # last year's dates (whose weekday coincidentally matches, so the
            # weekday snap above leaves it alone) shows up here as a backward
            # jump or a multi-day gap. Rebuild any such column from the running
            # sequence (previous day + 1) so the whole row is one continuous
            # range and can't leak junk records into a different month/year.
            ordered_cols = sorted(date_cols.keys())
            if len(ordered_cols) >= 2:
                contiguity_fixes = 0
                prev_date = None
                for c in ordered_cols:
                    d = date_cols[c]
                    if prev_date is not None:
                        delta = (d - prev_date).days
                        if delta < 1 or delta > 2:
                            d = prev_date + timedelta(days=1)
                            date_cols[c] = d
                            contiguity_fixes += 1
                    prev_date = date_cols[c]
                if contiguity_fixes:
                    logger.info("Excel import '%s': rebuilt %d non-contiguous date column(s) (%s .. %s)",
                                sheet.title, contiguity_fixes, date_cols[ordered_cols[0]], date_cols[ordered_cols[-1]])

            # Parse employee rows
            current_emp_code = None
            current_emp_name = None
            current_emp_hidden = False

            for r in range(start_data_row, num_rows + 1):
                emp_code_val = sheet_values[r-1][emp_code_col_idx - 1]
                emp_name_val = sheet_values[r-1][emp_name_col_idx - 1]
                branch_val = sheet_values[r-1][branch_col_idx - 1] if branch_col_idx else None

                row_has_identity = emp_code_val is not None or emp_name_val is not None
                if row_has_identity:
                    # A new employee block starts on the row carrying the name/code
                    # (the "In Time" row). HR marks people who have LEFT by HIDING
                    # their rows in Excel — they're invisible in the sheet but still
                    # physically present, and openpyxl reads them. Skip a hidden
                    # block entirely so ex-employees are never imported. The flag
                    # persists across the block's other rows (Out/Actual/Present),
                    # which don't re-trigger this detection.
                    current_emp_hidden = bool(sheet.row_dimensions[r].hidden)

                    current_emp_code = str(emp_code_val).strip() if emp_code_val is not None else None
                    if current_emp_code and current_emp_code.endswith(".0"):
                        current_emp_code = current_emp_code[:-2]
                    current_emp_name = str(emp_name_val).strip() if emp_name_val is not None else None

                    # Look up an existing employee's name when a row carries only a
                    # code. Match on the emp_code field within this sheet's branch
                    # (NOT on Employee.id — the code is not the primary key).
                    if not current_emp_hidden and not current_emp_name and current_emp_code:
                        emp_obj = Employee.objects.filter(
                            emp_code=str(current_emp_code), branch=sheet_branch
                        ).first()
                        if emp_obj:
                            current_emp_name = emp_obj.employee_name

                if current_emp_hidden or not current_emp_name:
                    continue

                row_type = normalize_val(sheet_values[r-1][type_col_idx - 1])

                # These sheets pre-print the 4-row type template (In/Out/Actual/
                # Present) far below the last real employee, as empty slots for
                # future hires. Every real employee block STARTS on an "In Time"
                # row that carries the name/code; an "In Time" row with no identity
                # of its own is therefore one of those blank template rows. Without
                # this guard those trailing rows stay attributed to the LAST real
                # employee and overwrite their captured punches with blanks — which
                # is exactly why the last person per sheet (e.g. Agalia in Vellore)
                # came out all-Absent. Detach from the previous employee here.
                if "in time" in row_type and not row_has_identity:
                    current_emp_name = None
                    continue

                if not row_type:
                    continue

                branch_norm = None
                if branch_val is not None:
                    bv_clean = " ".join(str(branch_val).split()).strip().lower()
                    if "chennai" in bv_clean:
                        branch_norm = "Chennai"
                    elif "vellore" in bv_clean:
                        branch_norm = "Vellore"
                    elif "salem" in bv_clean:
                        branch_norm = "Salem"
                    elif "kanchipuram" in bv_clean or "kanchi" in bv_clean:
                        branch_norm = "Kanchipuram"
                    elif "hosur" in bv_clean:
                        branch_norm = "Hosur"

                if not branch_norm:
                    branch_norm = sheet_branch

                # Identity key: (branch, NAME). The emp_code column in these
                # sheets is NOT unique — different people reuse the same code
                # (e.g. in Chennai both THAMARAIKKANNAN and SRI RAM N M are "2",
                # AKAASH and CHEZHIYEAN both "8"). Keying on code therefore MERGES
                # two distinct people into one row and cross-assigns their
                # attendance. Names ARE distinct within a branch, so the name is
                # the reliable identity; the code is retained only as metadata.
                # (current_emp_name is guaranteed non-empty here — rows without a
                # resolvable name are skipped above.)
                name_key = " ".join(current_emp_name.split()).lower()
                emp_key = (branch_norm, name_key)
                if emp_key not in employee_data:
                    employee_data[emp_key] = {
                        "code": current_emp_code,
                        "name": current_emp_name,
                        "branch": branch_norm,
                        "dates": {}
                    }
                elif branch_norm:
                    employee_data[emp_key]["branch"] = branch_norm

                dates_dict = employee_data[emp_key]["dates"]
                for c, date_obj in date_cols.items():
                    if date_obj not in dates_dict:
                        dates_dict[date_obj] = {"in": None, "out": None, "present": 0.0}

                    cell_val = sheet_values[r-1][c - 1]
                    # Never let a blank cell erase a value already captured for
                    # this date (defends against duplicate/stray rows for a person).
                    if "in time" in row_type:
                        if cell_val is not None:
                            dates_dict[date_obj]["in"] = cell_val
                    elif "out time" in row_type:
                        if cell_val is not None:
                            dates_dict[date_obj]["out"] = cell_val
                    elif "present" in row_type:
                        if cell_val is not None:
                            try:
                                dates_dict[date_obj]["present"] = float(cell_val)
                            except ValueError:
                                dates_dict[date_obj]["present"] = 0.0

        # 6. Database Transaction for bulk creation/updates
        employees_created = 0
        records_saved = 0

        try:
            with transaction.atomic():
                # Pre-fetch employees and user information to avoid repeated queries
                all_employees = list(Employee.objects.all().select_related("user"))
                # Branch-scoped lookup. Identity is (branch, NAME): names are
                # distinct within a branch while emp_code is not (people reuse
                # codes), so the name is authoritative. Names are normalized by
                # collapsing internal whitespace + lowercasing so " SRI  RAM "
                # matches "SRI RAM".
                def norm_name(s):
                    return " ".join(str(s or "").split()).lower()
                emp_by_branch_name = {(e.branch.lower().strip(), norm_name(e.employee_name)): e
                                      for e in all_employees}

                # Pre-fetch existing usernames to check uniqueness locally
                existing_usernames = set(User.objects.values_list("username", flat=True))

                # Pre-fetch existing attendance records within the imported date range
                all_dates = []
                for emp_info in employee_data.values():
                    all_dates.extend(list(emp_info["dates"].keys()))
                all_dates = list(set(all_dates))

                attendance_map = {}
                if all_dates:
                    min_date = min(all_dates)
                    max_date = max(all_dates)
                    start_datetime = make_aware(datetime.combine(min_date, time.min))
                    end_datetime = make_aware(datetime.combine(max_date, time.max))
                    existing_attendance = Attendance.objects.filter(
                        intime__range=(start_datetime, end_datetime)
                    )
                    for att in existing_attendance:
                        if att.employee_id and att.intime:
                            # Key on the LOCAL (IST) date. Absent records are stored
                            # at local midnight, which is the previous day in UTC, so
                            # att.intime.date() (UTC) would be off by one and cause
                            # duplicate rows on re-import. localtime() realigns it
                            # with the date_obj used when records are created below.
                            local_date = timezone.localtime(att.intime).date()
                            attendance_map[(att.employee_id, local_date)] = att

                records_to_create = []
                records_to_update = []
                # Records created THIS run, keyed by (employee_id, date). Prevents
                # duplicate rows when two sheet identities resolve to one employee
                # in the same import (attendance_map only holds pre-existing rows,
                # so without this a second identity wouldn't see the first's
                # just-created row and would insert a duplicate for that date).
                pending_map = {}
                updated_ids = set()

                for emp_key, emp_info in employee_data.items():
                    code = emp_info["code"]
                    name = emp_info["name"]
                    branch_val = emp_info.get("branch") or "Chennai"
                    dates = emp_info["dates"]

                    employee = None
                    bkey = branch_val.lower().strip()
                    # Match ONLY by (branch, NAME) — the authoritative identity,
                    # scoped to the branch so the same name in two branches
                    # (e.g. "PERUMAL" in both Salem and Kanchipuram) stays two
                    # people. emp_code is deliberately NOT used for matching:
                    # codes collide across people (a visible new joiner reuses a
                    # departed employee's code), so any code-based match
                    # cross-assigns one person's attendance to another
                    # (e.g. Vijayanand's punches landing on SRINI). An unmatched
                    # name creates a fresh employee below rather than borrowing a
                    # colliding code.
                    if name:
                        employee = emp_by_branch_name.get((bkey, norm_name(name)))

                    if not employee:
                        # 1. Create User account first
                        username = f"{name.lower().replace(' ', '')}_{code}" if code else name.lower().replace(' ', '')
                        base_username = username
                        counter = 1
                        while username in existing_usernames:
                            username = f"{base_username}_{counter}"
                            counter += 1

                        existing_usernames.add(username)

                        import secrets
                        clean_first_name = re.sub(r'[^a-zA-Z]', '', name.split()[0]).capitalize() if name else "User"
                        temp_pwd = f"{clean_first_name}@{secrets.randbelow(9000) + 1000}"

                        user_obj = User.objects.create_user(
                            username=username,
                            email=f"{username}@company.com",
                            first_name=name,
                            role='employee',
                            password=temp_pwd,
                            plain_password=temp_pwd
                        )

                        # 2. Create Employee profile linked to the User
                        create_kwargs = {
                            "user": user_obj,
                            "employee_name": name,
                            "role": "Staff",
                            "department": "General",
                            "salary": 0.00,
                            "branch": branch_val,
                            "status": "active",
                            # Persist the per-branch code (do NOT use it as the PK).
                            "emp_code": str(code) if code else None,
                        }
                        employee = Employee.objects.create(**create_kwargs)
                        emp_by_branch_name[(employee.branch.lower().strip(),
                                            norm_name(employee.employee_name))] = employee
                        employees_created += 1
                    else:
                        # Update branch if specified in Excel and differs from employee's current branch
                        if emp_info.get("branch") and employee.branch != emp_info["branch"]:
                            employee.branch = emp_info["branch"]
                            employee.save(update_fields=["branch"])

                        # Store the sheet's code as reference metadata on a
                        # name-matched employee (NOT used for matching — codes
                        # collide across people).
                        if code and not employee.emp_code:
                            employee.emp_code = str(code)
                            employee.save(update_fields=["emp_code"])

                        # If employee exists but doesn't have a linked user account, create and link one
                        if not employee.user:
                            username = f"{name.lower().replace(' ', '')}_{code}" if code else name.lower().replace(' ', '')
                            base_username = username
                            counter = 1
                            while username in existing_usernames:
                                username = f"{base_username}_{counter}"
                                counter += 1

                            existing_usernames.add(username)

                            import secrets
                            clean_first_name = re.sub(r'[^a-zA-Z]', '', name.split()[0]).capitalize() if name else "User"
                            temp_pwd = f"{clean_first_name}@{secrets.randbelow(9000) + 1000}"

                            user_obj = User.objects.create_user(
                                username=username,
                                email=f"{username}@company.com",
                                first_name=name,
                                role='employee',
                                password=temp_pwd,
                                plain_password=temp_pwd
                            )

                            employee.user = user_obj
                            employee.save()

                    # Create/Update attendance records
                    for date_obj, vals in dates.items():
                        in_val = vals["in"]
                        out_val = vals["out"]
                        present_val = vals["present"]

                        in_time = parse_excel_time(in_val)
                        out_time = parse_excel_time(out_val)

                        # Sunday is the weekly off. Classification order:
                        #  1. A real biometric punch (in_time) always wins -> Present,
                        #     even on a Sunday (someone who actually worked the off-day).
                        #  2. No punch on a Sunday -> Leave (weekly off), regardless of
                        #     any payroll "present day" flag.
                        #  3. No punch on a weekday but flagged present in payroll ->
                        #     Present (paid, placeholder 9:00 in-time, no out).
                        #  4. Otherwise -> Absent.
                        is_sunday = date_obj.weekday() == 6

                        if in_time is not None:
                            status = "Present"
                            in_datetime = make_aware(datetime.combine(date_obj, in_time))
                            out_datetime = make_aware(datetime.combine(date_obj, out_time)) if out_time else None
                        elif is_sunday:
                            status = "Leave"
                            in_datetime = make_aware(datetime.combine(date_obj, time.min))
                            out_datetime = None
                        elif present_val > 0.0:
                            status = "Present"
                            in_datetime = make_aware(datetime.combine(date_obj, time(9, 0)))
                            out_datetime = None
                        else:
                            status = "Absent"
                            in_datetime = make_aware(datetime.combine(date_obj, time.min))
                            out_datetime = None

                        attendance_fields = {
                            "employee": employee,
                            "employee_name": employee.employee_name,
                            "role": employee.role,
                            "department": employee.department,
                            "salary": employee.salary,
                            "intime": in_datetime,
                            "outtime": out_datetime,
                            "status": status
                        }

                        key = (employee.id, date_obj)
                        existing = attendance_map.get(key) or pending_map.get(key)
                        if existing:
                            for k, v in attendance_fields.items():
                                setattr(existing, k, v)
                            # Only queue DB rows for update; a pending create object
                            # is mutated in place and stays in records_to_create.
                            if existing.pk and id(existing) not in updated_ids:
                                records_to_update.append(existing)
                                updated_ids.add(id(existing))
                        else:
                            obj = Attendance(**attendance_fields)
                            records_to_create.append(obj)
                            pending_map[key] = obj
                        records_saved += 1

                # Execute bulk operations
                if records_to_create:
                    Attendance.objects.bulk_create(records_to_create)
                if records_to_update:
                    Attendance.objects.bulk_update(
                        records_to_update,
                        fields=["employee_name", "role", "department", "salary", "intime", "outtime", "status"]
                    )

        except Exception as e:
            return Response({"detail": f"Database transaction failed: {str(e)}"}, status=400)

        return Response({
            "message": f"Successfully processed {len(employee_data)} employees.",
            "employees_created": employees_created,
            "records_saved": records_saved
        }, status=200)

    @action(detail=False, methods=["delete"])
    def delete_all(self, request):
        user = request.user
        role = "superadmin" if user.is_superuser else getattr(user, 'role', 'employee')
        if role not in ["superadmin", "admin"]:
            return Response({"detail": "Permission denied."}, status=403)
        
        count, _ = Attendance.objects.all().delete()
        return Response({"detail": f"Successfully deleted {count} attendance records."}, status=200)

    def get_queryset(self):
        user = self.request.user
        queryset = Attendance.objects.select_related("employee").all().order_by("-id")
        role = "superadmin" if user.is_superuser else user.role
        if role == "employee":
            queryset = queryset.filter(employee__user=user)
        else:
            branches = get_allowed_branches(user, "attendance")
            if "All" not in branches:
                queryset = queryset.filter(employee__branch__in=branches)

        # Optional, non-breaking filters on `intime`. When none are supplied the
        # full list is returned exactly as before. The frontend opts in to a
        # smaller payload by sending any of these query params. Malformed values
        # are ignored rather than 500-ing.
        params = self.request.query_params

        # ?month=YYYY-MM  -> single calendar month
        month = params.get("month")
        if month:
            parts = month.split("-")
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                queryset = queryset.filter(
                    intime__year=int(parts[0]), intime__month=int(parts[1])
                )

        # ?start_date=YYYY-MM-DD and/or ?end_date=YYYY-MM-DD  -> inclusive range
        start_date = parse_date(params.get("start_date") or "")
        if start_date:
            queryset = queryset.filter(intime__date__gte=start_date)
        end_date = parse_date(params.get("end_date") or "")
        if end_date:
            queryset = queryset.filter(intime__date__lte=end_date)

        # ?employee=<id>
        employee_id = params.get("employee")
        if employee_id and employee_id.isdigit():
            queryset = queryset.filter(employee_id=int(employee_id))

        # ?status=Present
        status = params.get("status")
        if status:
            queryset = queryset.filter(status=status)

        return queryset

    @action(detail=False, methods=["post"])
    def check_in(self, request):
        user = request.user
        try:
            employee = user.employee_profile
        except AttributeError:
            return Response({"detail": "Employee profile not found for this user."}, status=400)

        # Fetch dynamically configured work location per user
        allowed_lat = employee.work_lat
        allowed_lon = employee.work_lon

        if allowed_lat is None or allowed_lon is None:
             return Response({"detail": "Allowed location not set for this employee. Please contact HR."}, status=400)

        user_lat = request.data.get("latitude")
        user_lon = request.data.get("longitude")

        if user_lat is None or user_lon is None:
            return Response({"detail": "Latitude and Longitude are required."}, status=400)

        try:
            distance = haversine_distance(float(user_lat), float(user_lon), float(allowed_lat), float(allowed_lon))
        except ValueError:
            return Response({"detail": "Invalid coordinates provided."}, status=400)

        if distance > 50:
            return Response({
                "detail": f"You are too far from the office ({round(distance)}m). Must be within 50m."
            }, status=403)

        # Check if already checked in today
        today = timezone.now().date()
        existing = Attendance.objects.filter(employee=employee, intime__date=today).first()
        
        if existing:
             return Response({"detail": "Already checked in for today."}, status=400)

        # Create record
        attendance = Attendance.objects.create(
            employee=employee,
            employee_name=employee.employee_name,
            role=employee.role,
            department=employee.department,
            salary=employee.salary,
            intime=timezone.now(),
            status="Present"
        )
        serializer = self.get_serializer(attendance)
        return Response(serializer.data, status=201)

    @action(detail=False, methods=["post"])
    def check_out(self, request):
        user = request.user
        try:
            employee = user.employee_profile
        except AttributeError:
            return Response({"detail": "Employee profile not found for this user."}, status=400)

        allowed_lat = employee.work_lat
        allowed_lon = employee.work_lon

        if allowed_lat is None or allowed_lon is None:
             return Response({"detail": "Allowed location not set for this employee. Please contact HR."}, status=400)

        user_lat = request.data.get("latitude")
        user_lon = request.data.get("longitude")

        if user_lat is None or user_lon is None:
            return Response({"detail": "Latitude and Longitude are required."}, status=400)

        try:
            distance = haversine_distance(float(user_lat), float(user_lon), float(allowed_lat), float(allowed_lon))
        except ValueError:
            return Response({"detail": "Invalid coordinates provided."}, status=400)

        if distance > 50:
            return Response({
                "detail": f"You are too far from the office ({round(distance)}m). Must be within 50m to clock out."
            }, status=403)

        today = timezone.now().date()
        existing = Attendance.objects.filter(employee=employee, intime__date=today).first()
        
        if not existing:
             return Response({"detail": "No clock-in found for today. Cannot clock out."}, status=400)

        if existing.outtime:
             return Response({"detail": "Already clocked out for today."}, status=400)

        existing.outtime = timezone.now()
        existing.save()
        serializer = self.get_serializer(existing)
        return Response(serializer.data, status=200)


class LeaveRequestViewSet(viewsets.ModelViewSet):
    serializer_class = LeaveRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = LeaveRequest.objects.select_related("employee").all().order_by("-applied_on")
        role = "superadmin" if user.is_superuser else user.role
        if role == "employee":
            return queryset.filter(employee__user=user)
        branches = get_allowed_branches(user, "leaves")
        if "All" not in branches:
            queryset = queryset.filter(employee__branch__in=branches)
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        employee = getattr(user, "employee_profile", None)
        if employee:
            serializer.save(employee=employee)
        else:
            serializer.save()

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        user = request.user
        role = "superadmin" if user.is_superuser else user.role
        if role not in ["superadmin", "admin", "hr"]:
            return Response({"detail": "Permission denied."}, status=403)
        leave = self.get_object()
        leave.status = "Approved"
        leave.save()
        return Response({"status": "Approved"})

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        user = request.user
        role = "superadmin" if user.is_superuser else user.role
        if role not in ["superadmin", "admin", "hr"]:
            return Response({"detail": "Permission denied."}, status=403)
        leave = self.get_object()
        leave.status = "Rejected"
        leave.save()
        return Response({"status": "Rejected"})

    
